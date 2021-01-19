import time

from conf import id_lawyer_dict
from lawyer import info_auth
import openpyxl

from log import logger

file_excel = '1.xlsx'

url_info = 'http://www.64365.com/lawyer/%s/info/'

result_title_list = ['name', 'location', 'company', 'skill', 'question',
                     'like', 'comment', 'score', 'phone_number', 'txt',
                     'middle_name', 'middle_detail', 'contact_phone_number', 'contact_wechat',
                     'contact_email', 'contact_address', 'more_policy_identity',
                     'more_job', 'more_traffic', 'more_hobby', 'career_office_name',
                     'career_size', 'career_position', 'career_number', 'career_desc',
                     'edu_education', 'edu_school', 'edu_speciality', 'person_gender', 'person_age',
                     'person_nation', 'person_zodiac', 'person_constellation', 'person_hometown',
                     'person_family', 'person_language']

result_title_cn_list = [
    '姓名', '省份', '公司', '技能', '解答咨询', '获得点赞', '获得好评', '综合得分', '联系电话', 'slogan', '姓名', '详细介绍', '联系方式-联系电话', '联系方式-微信',
    '联系方式-邮箱', '联系方式-联系地址', '更多信息-政治派别', '更多信息-政治职务', '更多信息-交通工具', '更多信息-兴趣爱好', '执业信息-执业律所', '执业信息-律所规模', '执业信息-职务',
    '执业信息-执业证号', '执业信息-团队描述', '教育背景-最高学历', '教育背景-毕业院校', '教育背景-所学专业', '个人资料-性别', '个人资料-年龄', '个人资料-民族', '个人资料-属相',
    '个人资料-星座', '个人资料-故乡', '个人资料-家庭', '个人资料-语言'
]

result_list = []
result_list.append(result_title_cn_list)


def to_excel(data_list):
    data_list = list(data_list)

    wb = openpyxl.Workbook()
    ws = wb.active

    for row_index, row_list in enumerate(data_list):
        row_index = row_index + 1

        for col_index, cell_value in enumerate(row_list):
            col_index = col_index + 1

            # logger.info(str(cell_value))

            value_result = str(cell_value)[:15000]


            try:
                ws.cell(row=row_index, column=col_index, value=value_result)
            except openpyxl.utils.exceptions.IllegalCharacterError:
                ws.cell(row=row_index, column=col_index, value='')

    wb.save(file_excel)


def main():
    # for k, v in id_lawyer_dict.items():
    #     time.sleep(2)
    #     url = url_info % str(k)
    #     logger.info(url)
    #     row_list = info_auth(url)
    #
    #     if any(row_list):
    #         result_list.append(row_list)

    result_list = []
    result_list.append(result_title_cn_list)
    with open('result.txt') as f:
        for i in f:
            i_new_list = eval(i)
            if any(i_new_list):
                result_list.append(i_new_list)

    to_excel(result_list)


if __name__ == '__main__':
    main()
